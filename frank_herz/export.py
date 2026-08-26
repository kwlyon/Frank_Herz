"""Excel export for complete dual-channel acquisition datasets."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import Calibration, DataPoint


DATA_HEADERS = (
    "Drive Voltage (V)",
    "Tube Current (pA)",
    "Elapsed Time (ms)",
    "Drive ADC Voltage (V)",
    "Current ADC Voltage (V)",
    "Drive ADC Raw (counts)",
    "Current ADC Raw (counts)",
    "Drive ADC Range (±V)",
    "Current ADC Range (±V)",
)


def export_xlsx(
    path: str | Path,
    points: Iterable[DataPoint],
    calibration: Calibration,
) -> int:
    """Write the complete dataset and calibration metadata to an .xlsx file."""

    rows = tuple(points)
    destination = Path(path)
    if destination.suffix.lower() != ".xlsx":
        destination = destination.with_suffix(".xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Acquisition Data"
    sheet.append(DATA_HEADERS)

    for point in rows:
        sheet.append(
            (
                point.drive_voltage_v,
                point.tube_current_pa,
                point.elapsed_ms,
                point.drive_adc_v,
                point.current_adc_v,
                point.drive_adc_counts,
                point.current_adc_counts,
                point.drive_adc_range_v,
                point.current_adc_range_v,
            )
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"
    widths = (20, 19, 19, 23, 25, 24, 26, 22, 24)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    metadata = workbook.create_sheet("Calibration")
    metadata.append(("Setting", "Value", "Units / meaning"))
    metadata_rows = (
        ("Drive voltage scale", calibration.drive_scale, "V actual per V measured"),
        ("Drive voltage offset", calibration.drive_offset_v, "V"),
        ("Picoammeter calibration", calibration.picoammeter_mv_per_pa, "mV per pA"),
        ("Picoammeter zero", calibration.picoammeter_zero_v, "V"),
        ("Picoammeter polarity", calibration.picoammeter_polarity, "+1 or -1"),
    )
    for row in metadata_rows:
        metadata.append(row)
    for cell in metadata[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    metadata.freeze_panes = "A2"
    metadata.column_dimensions["A"].width = 26
    metadata.column_dimensions["B"].width = 18
    metadata.column_dimensions["C"].width = 30

    workbook.save(destination)
    return len(rows)
